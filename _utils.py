import datetime
import glob
import os
import pickle
import re
import time

import math
import numpy as np
import torch
from loguru import logger
from openpyxl import Workbook  # pip install openpyxl
from openpyxl.reader.excel import load_workbook
from torchvision import transforms

from _data import get_topk, build_loader, get_class_num


def init(gpu_id: str):
    # os.environ["CUDA_VISIBLE_DEVICES"] = gpu_id  # before using torch
    torch.cuda.set_device(f"cuda:{gpu_id}")
    assert torch.cuda.is_available(), "CUDA is not available"
    # batch_size is too small will cause "Too many open files..."
    torch.multiprocessing.set_sharing_strategy("file_system")
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def find_diff_same(t1, t2, dim):
    """
    Args:
        t1 = torch.tensor([1, 9, 12, 5, 24])
        t2 = torch.tensor([1, 24])
    Returns:
        diff: torch.tensor([5, 9, 12])
        same: torch.tensor([1, 24])
    From:
        https://stackoverflow.com/questions/55110047
    """
    t1 = torch.unique(t1, dim=dim)
    t2 = torch.unique(t2, dim=dim)
    combined = torch.cat((t1, t2))
    uniques, counts = combined.unique(dim=dim, return_counts=True)
    difference = uniques[counts == 1]
    intersection = uniques[counts > 1]
    return difference, intersection


def calc_net_params(net, *args):
    n_parameters = sum(p.numel() for p in net.parameters() if p.requires_grad)
    for x in args:
        n_parameters += sum(p.numel() for p in x.parameters() if p.requires_grad)
    for n, p in net.named_parameters():
        if p.requires_grad:
            print(f"  {n}: {p.numel()} params")
    return n_parameters


def sizeof_fmt(num, suffix="B"):
    for unit in ("", "Ki", "Mi", "Gi", "Ti", "Pi", "Ei", "Zi"):
        if abs(num) < 1024.0:
            return f"{num:3.1f}{unit}{suffix}"
        num /= 1024.0
    return f"{num:.1f}Yi{suffix}"


def hash_center_type(n_classes, n_bits):
    """
    used in CenterHashing, CSQ, ...
    """
    lg2 = 0 if n_bits < 1 else int(math.log(n_bits, 2))
    if 2**lg2 != n_bits:
        return "random"

    if n_classes <= n_bits:
        return "ha_d"
    elif n_classes > n_bits and n_classes <= 2 * n_bits:
        return "ha_2d"
    else:
        return "random"


def prediction(net, dataloader, out_idx=-1, use_sign=True):
    device = next(net.parameters()).device
    codes, clses = [], []
    net.eval()
    logger.info(f"predicting({len(dataloader.dataset)})...")
    total_infer_time = 0.0  # 推理总时间初始化
    # for img, cls, _ in tqdm(dataloader):
    for x in dataloader:
        with torch.no_grad():
            start_time = time.time()  # 单次推理开始
            out = net(x[0].to(device))
            end_time = time.time()  # 单次推理结束

        total_infer_time += end_time - start_time  # 累加推理时间

        codes.append(out if out_idx == -1 else out[out_idx])
        clses.append(x[1])

        logger.info(f"Total inference time: {total_infer_time * 1000:.4f} seconds")

    return torch.cat(codes).sign() if use_sign else torch.cat(codes), torch.cat(clses).to(device)


def mean_average_precision(qB, rB, qL, rL, topk=-1) -> float:
    """
    Calculate mean average precision(map).

    Args:
        qB (torch.Tensor): Query data hash code.
        rB (torch.Tensor): Database data hash code.
        qL (torch.Tensor): Query data targets, one-hot
        rL (torch.Tensor): Database data targets, one-hot
        topk (float): Calculate top k data mAP.

    Returns:
        meanAP (float): Mean Average Precision.
    """
    num_query = qL.shape[0]
    if topk == -1:
        topk = rL.shape[0]
    mean_AP = 0.0
    for i in range(num_query):
        # Retrieve images from database
        retrieval = (qL[i, :] @ rL.T > 0).float()
        # Calculate hamming distance
        hamming_dist = 0.5 * (rB.shape[1] - qB[i, :] @ rB.T)
        # Arrange position according to hamming distance
        retrieval = retrieval[torch.argsort(hamming_dist)][:topk]
        # Retrieval count
        retrieval_cnt = retrieval.sum().int().item()
        # Can not retrieve images
        if retrieval_cnt == 0:
            continue
        # Generate score for every position
        score = torch.linspace(1, retrieval_cnt, retrieval_cnt).to(retrieval.device)
        # Acquire index
        index = ((retrieval == 1).nonzero(as_tuple=False).squeeze() + 1.0).float()
        mean_AP += (score / index).mean()
    mean_AP = mean_AP / num_query
    return mean_AP


def calc_hamming_dist(B1, B2):
    """
    calc Hamming distance
    Args:
        B1 (torch.Tensor): each bit of B1 ∈ {-1, 1}^k
        B2 (torch.Tensor): each bit of B2 ∈ {-1, 1}^k
    Returns:
        Hamming distance ∈ {0, k}
    """
    k = B2.shape[1]
    if len(B1.shape) < 2:
        B1 = B1.unsqueeze(0)
    distH = 0.5 * (k - B1.mm(B2.t()))
    return distH


def pr_curve(qB, rB, query_label, retrieval_label):
    num_query = qB.shape[0]
    num_bit = qB.shape[1]
    P = torch.zeros(num_query, num_bit + 1)
    R = torch.zeros(num_query, num_bit + 1)
    for i in range(num_query):
        gnd = (query_label[i].unsqueeze(0).mm(retrieval_label.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[i, :], rB)
        tmp = (hamm <= torch.arange(0, num_bit + 1).reshape(-1, 1).float().to(hamm.device)).float()
        total = tmp.sum(dim=-1)
        total = total + (total == 0).float() * 0.1
        t = gnd * tmp
        count = t.sum(dim=-1)
        p = count / total
        r = count / tsum
        P[i] = p
        R[i] = r
    mask = (P > 0).float().sum(dim=0)
    mask = mask + (mask == 0).float() * 0.1
    P = P.sum(dim=0) / mask
    R = R.sum(dim=0) / mask
    return P, R


def p_topK(qB, rB, qL, rL, K=None):
    if K is None:
        K = [1, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    num_query = qL.shape[0]
    p = [0] * len(K)
    for iter in range(num_query):
        gnd = (qL[iter].unsqueeze(0).mm(rL.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[iter, :], rB).squeeze()
        for i in range(len(K)):
            total = min(K[i], rL.shape[0])
            ind = torch.sort(hamm)[1][:total]
            gnd_ = gnd[ind]
            p[i] += gnd_.sum() / total
    p = torch.Tensor(p) / num_query
    return p


def cos(A, B=None):
    """cosine"""
    # An = normalize(A, norm='l2', axis=1)
    An = A / np.linalg.norm(A, ord=2, axis=1)[:, np.newaxis]
    if (B is None) or (B is A):
        return np.dot(An, An.T)
    # Bn = normalize(B, norm='l2', axis=1)
    Bn = B / np.linalg.norm(B, ord=2, axis=1)[:, np.newaxis]
    return np.dot(An, Bn.T)


def hamming(A, B=None):
    """A, B: [None, bit]
    elements in {-1, 1}
    """
    if B is None:
        B = A
    bit = A.shape[1]
    return (bit - A.dot(B.T)) // 2


def euclidean(A, B=None, sqrt=False):
    aTb = np.dot(A, B.T)
    if (B is None) or (B is A):
        aTa = np.diag(aTb)
        bTb = aTa
    else:
        aTa = np.diag(np.dot(A, A.T))
        bTb = np.diag(np.dot(B, B.T))
    D = aTa[:, np.newaxis] - 2.0 * aTb + bTb[np.newaxis, :]
    if sqrt:
        D = np.sqrt(D)
    return D


def NDCG(qF, rF, qL, rL, what=0, k=-1):
    """Normalized Discounted Cumulative Gain
    ref: https://github.com/kunhe/TALR/blob/master/%2Beval/NDCG.m
    """
    n_query = qF.shape[0]
    if (k < 0) or (k > rF.shape[0]):
        k = rF.shape[0]
    Rel = np.dot(qL, rL.T).astype(int)
    G = 2**Rel - 1
    D = np.log2(2 + np.arange(k))
    if what == 0:
        Rank = np.argsort(1 - cos(qF, rF))
    elif what == 1:
        Rank = np.argsort(hamming(qF, rF))
    elif what == 2:
        Rank = np.argsort(euclidean(qF, rF))

    _NDCG = 0
    for g, rnk in zip(G, Rank):
        dcg_best = (np.sort(g)[::-1][:k] / D).sum()
        if dcg_best > 0:
            dcg = (g[rnk[:k]] / D).sum()
            _NDCG += dcg / dcg_best
    return _NDCG / n_query


def get_precision_recall_by_Hamming_Radius(database_output, database_labels, query_output, query_labels, radius=2):
    bit_n = query_output.shape[1]
    ips = np.dot(query_output, database_output.T)
    ips = (bit_n - ips) / 2

    precX = []
    for i in range(ips.shape[0]):
        label = query_labels[i, :]
        label[label == 0] = -1
        idx = np.reshape(np.argwhere(ips[i, :] <= radius), (-1))
        all_num = len(idx)
        if all_num != 0:
            imatch = np.sum(database_labels[idx[:], :] == label, 1) > 0
            match_num = np.sum(imatch)
            precX.append(match_num / all_num)
        else:
            precX.append(0.0)
    return np.mean(np.array(precX))


class MyEval(object):
    main_dir, proj_name, backbone, dataset, n_bits = None, None, None, None, None

    def __init__(
        self,
        get_config_func,
        build_model_func,
        build_trans_func,
        build_loader_func,
        get_class_num_func,
        get_topk_func,
        proj_order,
        data_order,
        bits_order,
        the_suffix="",
    ):
        # functions for Supervised or ZeroShot Learning
        self.get_config = get_config_func
        self.build_model = build_model_func
        self.build_trans = build_trans_func
        self.build_loader = build_loader_func
        self.get_class_num = get_class_num_func
        self.get_topk = get_topk_func
        # Variables
        self.proj_order = proj_order
        self.data_order = data_order
        self.bits_order = bits_order
        self.the_suffix = the_suffix

    def prepare_excel(self, file_path, sheet_name):
        """
        Prepare the excel file to be written, create it if it doesn't exist.
        """
        is_new = False
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            is_new = True
            if "Sheet" in wb.sheetnames:
                ws = wb["Sheet"]
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        return wb, ws, is_new

    def get_ds_name(self, idx=-1):
        """
        Get the name of the database according to the different purposes.
        For example:
        1)self.dataset = "cifar"
          idx =  *: "cifar"
        2)self.dataset = (nus, voc)
          idx = -1: "nus->voc"
          idx =  0: "nus"
          idx =  1: "voc"
        """
        if isinstance(self.dataset, str):
            return self.dataset
        else:
            if idx == -1:
                return f"{self.dataset[0]}->{self.dataset[1]}"
            return self.dataset[idx]

    def write_excel_map(self, file_path, v):
        """
        support: NDCG, mAP, TrainingTime & EncodingTime
        """
        wb, ws, is_new = self.prepare_excel(file_path, self.proj_name)
        if is_new:
            for i, x in enumerate(self.bits_order["mAP"], 2):
                ws.cell(row=1, column=i).value = f"{x}bits"
            for i, x in enumerate(self.data_order, 2):
                ws.cell(row=i, column=1).value = x

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

        j = row1.index(f"{self.n_bits}bits") + 1
        i = col1.index(self.get_ds_name()) + 1

        ws.cell(row=i, column=j).value = v
        wb.save(file_path)

    def write_excel_hamming2(self, file_path, v):
        """
        save P@H≤2 result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, self.get_ds_name())
        if is_new:
            for i in range(len(self.proj_order)):
                ws.cell(row=1, column=i + 2).value = self.proj_order[i]
            for i, x in enumerate(self.bits_order["P@H≤2"], 2):
                ws.cell(row=i, column=1).value = f"{x}bits"

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

        j = row1.index(self.proj_name) + 1
        i = col1.index(f"{self.n_bits}bits") + 1

        ws.cell(row=i, column=j).value = v
        wb.save(file_path)

    def write_excel_pr(self, file_path, P, R):
        """
        save PR-curve result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, f"{self.get_ds_name()}@{self.n_bits}")
        if is_new:
            for i, x in enumerate(self.proj_order):
                ws.cell(row=1, column=2 * i + 1).value = x
                ws.cell(row=2, column=2 * i + 1).value = "R"
                ws.cell(row=2, column=2 * i + 2).value = "P"

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        try:
            j = row1.index(self.proj_name) + 1
        except ValueError:
            j = len(row1) + 1
            ws.cell(row=1, column=j).value = self.proj_name
            ws.cell(row=2, column=j).value = "R"
            ws.cell(row=2, column=j + 1).value = "P"

        for i, x in enumerate(R):
            ws.cell(row=i + 3, column=j).value = x.item()
            ws.cell(row=i + 3, column=j + 1).value = P[i].item()

        wb.save(file_path)

    def write_excel_topk(self, file_path, rst):
        """
        save TopN-precision result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, f"{self.get_ds_name()}@{self.n_bits}")

        j = 1
        if is_new:
            for i in range(len(self.proj_order)):
                ws.cell(row=1, column=i + 1).value = self.proj_order[i]
        else:
            while True:
                if ws.cell(row=1, column=j).value is None:
                    ws.cell(row=1, column=j).value = self.proj_name
                    break
                if ws.cell(row=1, column=j).value != self.proj_name:
                    j += 1
                    continue
                break

        for i in range(len(rst)):
            ws.cell(row=i + 2, column=j).value = rst[i].item()

        wb.save(file_path)

    def gen_cache_path(self):
        """
        Generate file path of the cache.
        """
        if isinstance(self.dataset, str):
            cache_path = f"{self.main_dir}/{self.proj_name}/output/{self.backbone}/{self.dataset}/{self.n_bits}/cache.p"
        else:
            if self.backbone is None:
                cache_path = (
                    f"{self.main_dir}/{self.proj_name}/output/{self.dataset[0]}_{self.dataset[1]}/{self.n_bits}/cache.p"
                )
            else:
                cache_path = f"{self.main_dir}/{self.proj_name}/output/{self.backbone}/{self.dataset[0]}/{self.n_bits}/{self.dataset[1]}_cache.p"
        return cache_path

    def get_checkpoint_path(self):
        """
        Get the file path of the checkpoint.
        """
        pkl_dir = os.path.dirname(self.gen_cache_path())
        pkl_list = glob.glob(f"{pkl_dir}/*.pkl")
        if len(pkl_list) != 1:
            logger.error(pkl_list)
            raise Exception(f"cannot locate one *.pkl in: {pkl_dir}")
        return pkl_list[0]

    def load_model(self):
        """
        Load pre-trained model based on configurations.
        """
        args = self.get_config()
        if isinstance(args, dict):
            args["backbone"] = self.backbone
            args["n_bits"] = self.n_bits
            if "n_classes" in args.keys():
                args["n_classes"] = self.get_class_num(self.get_ds_name(0))
        else:
            args.backbone = self.backbone
            args.n_bits = self.n_bits
            if "n_classes" in args:
                args.n_classes = self.get_class_num(self.get_ds_name(0))
        # build model in cuda
        out = self.build_model(args, pretrained=False)
        if isinstance(out, tuple):
            net, out_idx = out
        else:
            net = out
            out_idx = -1
        # load checkpoint
        checkpoint_path = self.get_checkpoint_path()
        checkpoint = torch.load(checkpoint_path, map_location="cpu")
        msg = net.load_state_dict(checkpoint, strict=False)
        logger.info(f"model loaded: {msg}")

        return net, out_idx

    def build_loaders(self):
        """
        Build query & retrieval data loaders based on configurations.
        """
        args = self.get_config()
        dataset = self.get_ds_name(1)
        data_dir = "../_datasets" + self.the_suffix
        if isinstance(args, dict):
            args["dataset"] = dataset
            args["data_dir"] = data_dir
        else:
            args.dataset = dataset
            args.data_dir = data_dir

        if self.build_trans is not None:
            out = self.build_trans(args)
            if isinstance(out, tuple):
                trans = out[1]
            else:
                trans = out
        else:
            logger.debug("use default transforms")
            trans = transforms.Compose(
                [
                    transforms.Resize(256),
                    transforms.CenterCrop(224),
                    transforms.ToTensor(),
                    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
                ]
            )
        query_loader = self.build_loader(data_dir, dataset, "query", trans, batch_size=100, num_workers=4)
        dbase_loader = self.build_loader(data_dir, dataset, "dbase", trans, batch_size=100, num_workers=4)

        return query_loader, dbase_loader

    def load_prediction(self):
        """
        Load these four values, with cache use cache, without cache recalculate.
        """
        cache_path = self.gen_cache_path()

        if not os.path.exists(cache_path):
            qB, qL, rB, rL = self.calc_prediction()
            save_obj = {
                "qB": qB.cpu(),
                "qL": qL.cpu(),
                "rB": rB.cpu(),
                "rL": rL.cpu(),
            }
            with open(cache_path, "ab") as f:
                pickle.dump(save_obj, f)
        else:
            logger.debug("load qB, qL, rB, rL from cache")
            with open(cache_path, "rb") as f:
                data = pickle.load(f)
            qB, qL, rB, rL = data["qB"].cuda(), data["qL"].cuda(), data["rB"].cuda(), data["rL"].cuda()
        return qB, qL, rB, rL

    def calc_prediction(self):
        """
        A pre-trained model is constructed to compute these 4 values.
        """
        net, out_idx = self.load_model()

        global query_loader
        global dbase_loader

        if "query_loader" in globals():
            if query_loader.dataset.name != self.get_ds_name(1):
                logger.debug(f"{query_loader.dataset.name} != {self.get_ds_name(1)}, loader need to be rebuilt")
                del query_loader
                del dbase_loader
            else:
                logger.debug(f"loader is the same, no need to be rebuilt")

        if "query_loader" not in globals():
            query_loader, dbase_loader = self.build_loaders()

        qB, qL = prediction(net, query_loader, out_idx)
        rB, rL = prediction(net, dbase_loader, out_idx)

        return qB, qL, rB, rL

    def get_training_time(self):
        """
        Get the training time of the best map through the log file.
        """
        log_path = os.path.dirname(self.gen_cache_path()) + "/train.log"

        with open(log_path, "r") as f:
            lines = f.read().splitlines()
        ret = []
        epoch = 0
        for line in lines:
            m = re.search(r".+\[Training].+\[epoch:([0-9./]+)]\[time:([0-9.]+)].+", line)
            if m:
                ret.append(m.group(2))
                epoch = int(m.group(1).split("/")[0])
                # print(epoch, m.group(2))
        ret = np.array(ret).astype(float)
        assert epoch != 0, f"can't find training time"
        assert len(ret) == epoch + 1, f"may missing some training time"

        idx = 0
        for line in lines[-5:]:
            # print(line)
            m = re.search(r".+best epoch: (\d+).+", line)
            if m:
                idx = int(m.group(1))
                # print(m.group(1))
                break
        assert idx != 0, "can't find best epoch"

        return ret[: idx + 1].sum()

    def get_encoding_time(self):
        """
        Get the encoding time of the model-generated hash based on the following parameters:
        dataset = "voc"
        batch_size = 100
        usage = "query"
        """
        trans = transforms.Compose(
            [
                transforms.Resize((224, 224)),
                transforms.ToTensor(),
                transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
            ]
        )
        dataloader = self.build_loader(
            "../_datasets_zs", "voc", "query", trans, batch_size=100, shuffle=False, num_workers=4
        )
        net, _ = self.load_model()
        net.eval()
        logger.info(f"predicting({len(dataloader.dataset)})...")
        tic = time.time()
        for x in dataloader:
            with torch.no_grad():
                net(x[0].cuda())
        toc = time.time()
        return toc - tic

    def get_run_list(self, evals, n_bits):
        """
        Check whether to run an evaluation based on the number of bits.
        """
        rst = []
        for x in evals:
            if n_bits in self.bits_order[x]:
                rst.append(x)
        return rst

    def __call__(
        self,
        main_dir,
        proj_name,
        backbone,
        evals,
        datasets,
        hash_bits,
        use_cache=True,
        save_xlsx=True,
    ):
        self.main_dir = main_dir
        self.proj_name = proj_name
        self.backbone = backbone
        for dataset in datasets:
            self.dataset = dataset
            logger.info(f"processing dataset: {self.get_ds_name()}")

            for hash_bit in hash_bits:
                self.n_bits = hash_bit
                logger.info(f"processing hash-bit: {hash_bit}")

                run_list = self.get_run_list(evals, hash_bit)

                if len(run_list) == 0:
                    logger.info(f"no eval to run, pass")
                    continue

                if self.get_ds_name(0) == self.get_ds_name(1) or proj_name in ["G_MLZSL", "T_MLZSH"]:
                    is_useful = True
                else:
                    is_useful = False

                if not (
                    (
                        not is_useful
                        and set(run_list) <= {"TrainingTime", "EncodingTime", "SelfCheck"}
                        or (is_useful and set(run_list) <= {"TrainingTime", "EncodingTime"})
                    )
                ):
                    if use_cache:
                        qB, qL, rB, rL = self.load_prediction()
                    else:
                        qB, qL, rB, rL = self.calc_prediction()

                # calc mAP
                if "mAP" in run_list:
                    topk = self.get_topk(self.get_ds_name(1))
                    map_k = mean_average_precision(qB, rB, qL, rL, topk)
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/eval_map{self.the_suffix}.xlsx", f"{map_k:.3f}")
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][mAP@{topk}:{map_k:.3f}]")
                else:
                    map_k = None

                # calc PR curve
                if "PR-curve" in run_list:
                    P, R = pr_curve(qB, rB, qL, rL)
                    if save_xlsx:
                        self.write_excel_pr(f"{main_dir}/eval_pr{self.the_suffix}.xlsx", P, R)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][PR-curve is done]")

                # calc TopN precision
                if "TopN-precision" in run_list:
                    rst = p_topK(qB, rB, qL, rL)
                    if save_xlsx:
                        self.write_excel_topk(f"{main_dir}/eval_topk{self.the_suffix}.xlsx", rst)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][TopN-precision is done]")

                if "NDCG" in run_list or "P@H≤2" in run_list:
                    qB = qB.cpu().numpy()
                    rB = rB.cpu().numpy()
                    qL = qL.cpu().numpy()
                    rL = rL.cpu().numpy()

                # calc NDCG
                if "NDCG" in run_list:
                    ndcg = NDCG(qB, rB, qL, rL, what=1, k=1000)
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/eval_ndcg{self.the_suffix}.xlsx", f"{ndcg:.3f}")
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][NDCG:{ndcg:.3f}]")

                # calc Precision curves within Hamming Radius 2
                if "P@H≤2" in run_list:
                    prec = get_precision_recall_by_Hamming_Radius(rB, rL, qB, qL)
                    if save_xlsx:
                        self.write_excel_hamming2(f"{main_dir}/eval_hamming2{self.the_suffix}.xlsx", prec)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][P@H≤2:{prec:.3f}]")

                if "TrainingTime" in run_list:
                    rst = self.get_training_time()
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/training_time{self.the_suffix}.xlsx", rst)
                    logger.info(
                        f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][TrainingTime:{datetime.timedelta(seconds=rst)}]"
                    )

                if "EncodingTime" in run_list:
                    rst = self.get_encoding_time()
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/encoding_time{self.the_suffix}.xlsx", rst)
                    logger.info(
                        f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][EncodingTime:{datetime.timedelta(seconds=rst)}]"
                    )

                if "SelfCheck" in run_list:
                    # get mAP from check pint file name
                    pkl_fn = os.path.basename(self.get_checkpoint_path())
                    m = re.search(r"e\d+_([0-9.]+)\.pkl", pkl_fn)
                    if m:
                        map_pkl = float(m.group(1))
                    else:
                        m = re.search(r"iter\d+_([0-9.]+)\.pkl", pkl_fn)
                        if m:
                            map_pkl = float(m.group(1))
                        else:
                            raise Exception(f"can't extract mAP from file name: {pkl_fn}")
                    # get mAP from model prediction
                    temp = None
                    if is_useful:
                        if map_k is None:
                            topk = self.get_topk(self.get_ds_name(1))
                            map_k = mean_average_precision(qB, rB, qL, rL, topk)
                    else:
                        topk = self.get_topk(self.get_ds_name(0))
                        temp = self.dataset
                        self.dataset = (self.dataset[0], self.dataset[0])
                        # logger.debug(self.dataset)
                        if use_cache:
                            qB, qL, rB, rL = self.load_prediction()
                        else:
                            qB, qL, rB, rL = self.calc_prediction()
                        map_k = mean_average_precision(qB, rB, qL, rL, topk)

                    map_k = round(map_k.item(), 3)

                    if map_k != map_pkl:
                        # logger.warning(f"map[{map_k}] != map_pkl[{map_pkl}]")
                        logger.warning(
                            f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][SelfCheck:failed][mAP@{topk}:{map_k}][mAP@pkl:{map_pkl}]"
                        )
                    else:
                        # logger.debug(f"map[{map_k}] = map_pkl[{map_pkl}]")
                        logger.info(
                            f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][SelfCheck:passed][mAP@{topk}:{map_k}][mAP@pkl:{map_pkl}]"
                        )
                    if temp is not None:
                        self.dataset = temp


def init_my_eval(get_config, build_model, build_trans):
    proj_order = ["DPSH", "DSH", "CSQ", "OrthoHash", "IDHN", "HyP2", "CenterHashing", "SWTH", "SPRCH"]
    data_order = ["cifar", "nuswide", "flickr", "coco"]
    bits_order = {
        "mAP": [16, 32, 64, 128],
        "NDCG": [16, 32, 64, 128],
        "PR-curve": [16, 32, 48, 64, 128],
        "TopN-precision": [16, 32, 48, 64, 128],
        "P@H≤2": [16, 32, 48, 64],
        "TrainingTime": [16, 32, 48, 64, 128],
        "EncodingTime": [16, 32, 48, 64, 128],
        "SelfCheck": [16, 32, 48, 64, 128],
    }
    my_eval = MyEval(
        get_config, build_model, build_trans, build_loader, get_class_num, get_topk, proj_order, data_order, bits_order
    )
    return my_eval


if __name__ == "__main__":
    t1 = torch.tensor([1, 9, 9, 12, 5, 24])
    t2 = torch.tensor([1, 24])
    d, s = find_diff_same(t1, t2, 0)
    print(d)
    print(s)
